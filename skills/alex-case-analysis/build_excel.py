#!/usr/bin/env python3
"""
tian-case-organizer — Extract Chinese court case info from converted Markdown
and generate formatted Excel.
Usage: python3 build_excel.py <_converted_dir> <output_xlsx_path>
"""
import sys
import os
import re
import glob
import html
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Chinese digit mapping ──────────────────────────────────────
CN_DIGITS = {
    '〇': 0, '零': 0, '一': 1, '二': 2, '三': 3, '四': 4,
    '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10,
}


def clean_html(s):
    """Strip HTML tags and unescape entities."""
    s = re.sub(r'<[^>]+>', '', s)
    s = html.unescape(s)
    return s.replace('&gt;', '>').replace('&lt;', '<').replace('&amp;', '&')


def normalize_date(raw):
    """Convert any date format to '2021年12月31日'."""
    if not raw:
        return ''
    raw = raw.strip()

    # Already target format?
    if re.match(r'^\d{4}年\d{1,2}月\d{1,2}日$', raw):
        return raw

    # 2021.12.31 or 2021.12.31 00:00:00
    m = re.match(r'(\d{4})[./年](\d{1,2})[./月](\d{1,2})日?', raw)
    if m:
        return f"{m.group(1)}年{int(m.group(2))}月{int(m.group(3))}日"

    # 二〇二一年十二月三十一日
    m = re.match(r'([二三四五六七八九〇零]{2,4})年([一二三四五六七八九十]+)月([一二三四五六七八九十]+)日', raw)
    if m:
        year = ''.join(str(CN_DIGITS.get(c, c)) for c in m.group(1))
        month = cn_num_to_int(m.group(2))
        day = cn_num_to_int(m.group(3))
        return f"{year}年{month}月{day}日"

    return raw


def cn_num_to_int(s):
    """Convert Chinese numeral string like '二十六' to integer 26."""
    if s in CN_DIGITS and CN_DIGITS[s] != 10:
        return CN_DIGITS[s]
    if s.startswith('十'):
        return 10 + (CN_DIGITS.get(s[1], 0) if len(s) > 1 else 0)
    if '十' in s:
        parts = s.split('十')
        tens = CN_DIGITS.get(parts[0], 1)
        ones = CN_DIGITS.get(parts[1], 0) if len(parts) > 1 and parts[1] else 0
        return tens * 10 + ones
    return sum(CN_DIGITS.get(c, 0) for c in s)


def simplify_procedure(raw):
    """民事一审→一审, 民事二审→二审, 民事再审→再审."""
    if not raw:
        return ''
    mapping = {'民事一审': '一审', '民事二审': '二审', '民事再审': '再审',
               '刑事一审': '一审', '刑事二审': '二审', '行政一审': '一审', '行政二审': '二审'}
    for k, v in mapping.items():
        if k in raw:
            return v
    return raw


def extract_case_name(text):
    """Extract case name from markdown title."""
    for line in text.split('\n')[:5]:
        line = clean_html(line).strip()
        line = re.sub(r'^#+\s*', '', line).strip()
        if line and len(line) > 10 and ('纠纷' in line or '案' in line):
            return line
    for line in text.split('\n')[:10]:
        line = clean_html(line).strip()
        line = re.sub(r'^#+\s*', '', line).strip()
        if line and len(line) > 15:
            return line
    return ''


def parse_date_for_sort(date_str):
    """Parse '2021年12月31日' into (2021, 12, 31) tuple for sorting. Returns (9999,0,0) if unparseable."""
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return (9999, 0, 0)


def extract_metadata(text):
    """Extract header metadata fields from case markdown."""
    meta = {}

    # ── 案号 ──
    # HTML table format: <td>案号:</td><td>(2022)京73民终555号</td>
    m = re.search(r'案号[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:600])
    if m:
        meta['案号'] = m.group(1).strip()
    if not meta.get('案号'):
        m = re.search(r'案号[：:\s]*(.{10,60}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = clean_html(m.group(1)).strip()
            if val and not val.startswith('<'):
                meta['案号'] = val
    if not meta.get('案号'):
        # Try body text: （2021）陕01知民初3078号
        m = re.search(r'[（(]\s*(\d{4})\s*[）)]\s*[^\n]{5,30}?\d+\s*号', text[:800])
        if m:
            meta['案号'] = m.group(0).strip()

    # ── 审理法院 ──
    m = re.search(r'审理法院[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:600])
    if m:
        meta['审理法院'] = m.group(1).strip()
    if not meta.get('审理法院'):
        m = re.search(r'审理法院[：:\s]*([^<\n]{2,40}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<') and not val.startswith('/'):
                meta['审理法院'] = clean_html(val)
    if not meta.get('审理法院'):
        m = re.search(r'(.{2,30}?(?:人民法院|法院))\s*\n\s*(?:民事|刑事|行政)(?:判决书|裁定书)', text[:500])
        if m:
            meta['审理法院'] = m.group(1).strip()

    # ── 裁判日期 ──
    m = re.search(r'审结日期[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:600])
    if m:
        meta['裁判日期'] = normalize_date(m.group(1).strip())
    if not meta.get('裁判日期'):
        m = re.search(r'审结日期[：:\s]*([^<\n]{4,20}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<'):
                meta['裁判日期'] = normalize_date(clean_html(val))
    if not meta.get('裁判日期'):
        m = re.search(r'裁判日期[：:\s]*([^<\n]{4,20}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<'):
                meta['裁判日期'] = normalize_date(clean_html(val))
    if not meta.get('裁判日期'):
        m = re.search(r'二[〇零一二三四五六七八九十]{3,4}年[一二三四五六七八九十]+月[一二三四五六七八九十]+日', text)
        if m:
            meta['裁判日期'] = normalize_date(m.group(0))

    # ── 审判程序 ──
    m = re.search(r'案件类型[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:600])
    if m:
        meta['审判程序'] = simplify_procedure(m.group(1).strip())
    if not meta.get('审判程序'):
        m = re.search(r'案件类型[：:\s]*([^<\n]{2,20}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<'):
                meta['审判程序'] = simplify_procedure(clean_html(val))
    if not meta.get('审判程序'):
        m = re.search(r'审理程序[：:\s]*([^<\n]{2,20}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<'):
                meta['审判程序'] = simplify_procedure(clean_html(val))
    if not meta.get('审判程序'):
        m = re.search(r'(民事一审|民事二审|民事再审)', text[:800])
        if m:
            meta['审判程序'] = simplify_procedure(m.group(1))

    # ── 原告/上诉人 ──
    if not meta.get('原告'):
        m = re.search(r'相关企业[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:600])
        if m:
            companies = m.group(1).strip().split()
            if len(companies) >= 2:
                meta['原告'] = companies[0]
                meta['被告'] = companies[1]
    if not meta.get('原告'):
        m = re.search(r'原告[：:]\s*([^\n。，,]{5,80}?)(?:[，,]\s*住所|。|\n)', text[:2000])
        if m:
            meta['原告'] = m.group(1).strip().rstrip('，,。；;')
    if not meta.get('原告'):
        m = re.search(r'上诉人[（(]原审[^）)]*[）)]?[：:]?\s*([^\n。，,]{5,80}?)(?:[，,]\s*住所|。|\n)', text[:2000])
        if m:
            meta['原告'] = m.group(1).strip().rstrip('，,。；;')
    # From judgment body: 原告XXX公司（以下简称XXX）
    if not meta.get('原告'):
        m = re.search(r'原告\s*([^\n（(]{5,60}?)[（(]以下简称', text[:2000])
        if m:
            meta['原告'] = m.group(1).strip()
    if not meta.get('原告'):
        m = re.search(r'上诉人\s*([^\n（(]{5,60}?)[（(]以下简称', text[:2000])
        if m:
            meta['原告'] = m.group(1).strip()

    # ── 被告/被上诉人 ──
    if not meta.get('被告'):
        m = re.search(r'被告[：:]\s*([^\n。，,]{5,80}?)(?:[，,]\s*住所|。|\n)', text[:2500])
        if m:
            meta['被告'] = m.group(1).strip().rstrip('，,。；;')
    if not meta.get('被告'):
        m = re.search(r'被上诉人[（(]原审[^）)]*[）)]?[：:]?\s*([^\n。，,]{5,80}?)(?:[，,]\s*住所|。|\n)', text[:2500])
        if m:
            meta['被告'] = m.group(1).strip().rstrip('，,。；;')
    if not meta.get('被告'):
        m = re.search(r'被告\s*([^\n（(]{5,60}?)[（(]以下简称', text[:3000])
        if m:
            meta['被告'] = m.group(1).strip()
    if not meta.get('被告'):
        m = re.search(r'被上诉人\s*([^\n（(]{5,60}?)[（(]以下简称', text[:3000])
        if m:
            meta['被告'] = m.group(1).strip()

    # ── 文书类型 ──
    m = re.search(r'文书类型[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:600])
    if m:
        meta['文书类型'] = m.group(1).strip()
    if not meta.get('文书类型'):
        m = re.search(r'文书类型[：:\s]*([^<\n]{2,10}?)(?:</t[dh]|\n)', text[:600])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<'):
                meta['文书类型'] = clean_html(val)
    if not meta.get('文书类型'):
        meta['文书类型'] = '判决书' if '判决书' in text[:200] else ('裁定书' if '裁定书' in text[:200] else '')

    # ── 来源 (header only, first 500 chars) ──
    m = re.search(r'来源[：:\s]*</td>\s*<td[^>]*>\s*([^<]+)', text[:500])
    if m:
        meta['来源'] = m.group(1).strip()
    if not meta.get('来源'):
        m = re.search(r'来源[：:\s]*([^<\n]{3,80}?)(?:</t[dh]|\n)', text[:500])
        if m:
            val = m.group(1).strip()
            if val and not val.startswith('<') and not val.startswith('/'):
                meta['来源'] = clean_html(val)
    if not meta.get('来源'):
        meta['来源'] = '北大法宝'

    return meta


def extract_court_opinion(text):
    """Extract court's reasoning section verbatim."""
    start_patterns = [
        r'\n本院认为[：:\s,，]*\n',
        r'本院认为[：:，,]',
        r'\n(?:法院生效)?裁判认为[：:\s]*\n',
        r'\n##\s*【?裁判理由】?\s*\n',
        r'\n##\s*【?裁判内容】?\s*\n',
        r'\n##\s*【?评析】?\s*\n',
    ]
    start_pos = None
    for pat in start_patterns:
        m = re.search(pat, text)
        if m:
            start_pos = m.start()
            body_start = m.end()
            break

    if start_pos is None:
        return ''

    end_patterns = [
        r'\n审\s*判\s*长\s', r'\n审判长\s',
        r'\n审\s*判\s*员\s', r'\n审判员\s',
        r'\n书\s*记\s*员\s', r'\n书记员\s',
        r'\n## 本案法律依据', r'\n© 北大法宝',
        r'\n## 【?裁判要旨】?', r'\n## 【?关联索引】?',
        r'\n## 【?典型意义】?', r'\n## 同案由重要案例',
        r'\n## 本法院同类案例',
    ]
    end_pos = len(text)
    for pat in end_patterns:
        em = re.search(pat, text[body_start:])
        if em:
            end_pos = body_start + em.start()
            break

    return text[body_start:end_pos].strip()


def extract_dispute_focus(text):
    """Extract dispute focus - complete text, no truncation."""
    patterns = [
        r'本案[的]?争议焦点[在于：:是，,\s]*(.+?)(?:[。；;]\s*(?:结合|本案|综上|因此|故|本?院))',
        r'争议焦点[：:]\s*(.+?)(?:。\s*\n)',
        r'本案争议焦点在于[：:是如何，,\s]*(.+?)(?:[。])',
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            return m.group(1).strip()
    return ''


def extract_judgment_result(text):
    """Extract judgment result."""
    m = re.search(r'判决如下[：:\s]*\n(.*?)(?=\n\s*(?:审判长|审\s*判|书记员|本案法律依据|©))', text, re.DOTALL)
    if m:
        return m.group(1).strip()[:600]
    m = re.search(r'裁判结果[：:]\s*(.+?)(?:\n\n|\n##|\n©|\Z)', text)
    if m:
        return m.group(1).strip()[:600]
    m = re.search(r'(?:遂判决|判决[：:]|裁定[：:])(.+?)(?:\n\n|\n#|\Z)', text)
    if m:
        return m.group(1).strip()[:600]
    return ''


def parse_case_file(filepath):
    """Parse a single converted .md file, return case data dict."""
    with open(filepath, 'r', encoding='utf-8') as f:
        text = f.read()

    meta = extract_metadata(text)
    case_name = extract_case_name(text)
    if not case_name:
        case_name = os.path.basename(os.path.dirname(filepath))[:60]

    return {
        '案件名称': case_name,
        '案号': meta.get('案号', ''),
        '审理法院': meta.get('审理法院', ''),
        '裁判日期': meta.get('裁判日期', ''),
        '审判程序': meta.get('审判程序', ''),
        '原告': meta.get('原告', ''),
        '被告': meta.get('被告', ''),
        '争议焦点': extract_dispute_focus(text),
        '裁判结果': extract_judgment_result(text),
        '文书类型': meta.get('文书类型', ''),
        '来源': meta.get('来源', '北大法宝'),
        '本院认为': extract_court_opinion(text),
    }


def deduplicate_cases(cases):
    """Merge cases with identical case numbers. Keep version with most complete data."""
    seen = {}
    result = []
    for case in cases:
        cn = case.get('案号', '')
        if cn and cn in seen:
            existing = seen[cn]
            for key in case:
                if not existing.get(key) and case.get(key):
                    existing[key] = case[key]
            continue
        if cn:
            seen[cn] = case
        result.append(case)
    return result


def sort_by_date(cases):
    """Sort cases by judgment date ascending. Cases without dates go last."""
    return sorted(cases, key=lambda c: parse_date_for_sort(c.get('裁判日期', '')))


# ── Excel Formatting Constants ──────────────────────────────────
HEADER_FONT = Font(name='黑体', size=11, bold=True)
BODY_FONT = Font(name='宋体', size=10)
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)


def create_excel(cases, output_path):
    """Create formatted Excel workbook with two sheets."""
    wb = Workbook()

    # ═══ Sheet 1: 案件基本信息 ═══
    ws1 = wb.active
    ws1.title = "案件基本信息"

    headers1 = [
        '序号', '案件名称', '案号', '审理法院', '裁判日期',
        '审判程序', '原告（上诉人）', '被告（被上诉人）',
        '争议焦点', '裁判结果', '文书类型', '来源', '备注'
    ]
    keys1 = [
        None, '案件名称', '案号', '审理法院', '裁判日期',
        '审判程序', '原告', '被告',
        '争议焦点', '裁判结果', '文书类型', '来源', None
    ]
    col_widths1 = [6, 35, 28, 22, 16, 10, 28, 28, 45, 50, 10, 20, 15]

    # Header row
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, case in enumerate(cases, 2):
        for col_idx, (header, key) in enumerate(zip(headers1, keys1), 1):
            if key is None and col_idx == 1:
                val = row_idx - 1
            elif key is None:
                val = ''
            else:
                val = case.get(key, '')
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER_WRAP
            cell.border = THIN_BORDER
        ws1.row_dimensions[row_idx].height = 100

    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'

    # ═══ Sheet 2: 本院认为（原文） ═══
    ws2 = wb.create_sheet(title="本院认为（原文）")

    headers2 = ['序号', '案件名称', '案号', '本院认为（原文全文）']
    col_widths2 = [6, 35, 28, 120]

    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_WRAP
        cell.border = THIN_BORDER

    for row_idx, case in enumerate(cases, 2):
        vals = [row_idx - 1, case.get('案件名称', ''), case.get('案号', ''), case.get('本院认为', '')]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER_WRAP
            cell.border = THIN_BORDER
        opinion_len = len(case.get('本院认为', ''))
        ws2.row_dimensions[row_idx].height = max(60, min(800, opinion_len * 0.06))

    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 build_excel.py <_converted_dir> <output_xlsx_path>")
        sys.exit(1)

    converted_dir = sys.argv[1]
    output_path = sys.argv[2]

    md_files = sorted(glob.glob(os.path.join(converted_dir, '*/full.md')))
    if not md_files:
        print("No converted markdown files found in:", converted_dir)
        sys.exit(1)

    print(f"Found {len(md_files)} converted case files.\n")

    cases = []
    for f in md_files:
        dirname = os.path.basename(os.path.dirname(f))
        print(f"  Processing: {dirname[:60]}...")
        try:
            data = parse_case_file(f)
            cases.append(data)
            has_op = '✓' if data.get('本院认为') else '✗'
            print(f"    {data['案件名称'][:45]} | 法院: {data.get('审理法院','?')[:15]} | 本院认为: {has_op}")
        except Exception as e:
            print(f"    ERROR: {e}")

    # Dedup
    before = len(cases)
    cases = deduplicate_cases(cases)
    dup_count = before - len(cases)
    if dup_count:
        print(f"\n  去重合并: {dup_count} 组重复案号")

    # Sort by date
    cases = sort_by_date(cases)

    # Generate Excel
    create_excel(cases, output_path)

    # Summary report
    dates = [c.get('裁判日期', '') for c in cases if c.get('裁判日期')]
    with_opinion = sum(1 for c in cases if c.get('本院认为'))

    print(f"\n{'='*60}")
    print(f"案件整理完成")
    print(f"{'─'*60}")
    print(f"扫描文件：{len(md_files)} 个 PDF")
    print(f"成功处理：{len(cases)} 个案件")
    print(f"去重合并：{dup_count} 组重复")
    print(f"本院认为：{with_opinion}/{len(cases)} 个案件有完整原文")
    if dates:
        print(f"日期跨度：{dates[0]} — {dates[-1]}")
    print(f"输出文件：{output_path}")
    print(f"{'─'*60}")

    for i, c in enumerate(cases, 1):
        op = f"✓ {len(c.get('本院认为', ''))}字" if c.get('本院认为') else '✗'
        print(f"  {i}. {c.get('案件名称', '?')[:40]} | {c.get('裁判日期', '无日期')} | {op}")


if __name__ == '__main__':
    main()
